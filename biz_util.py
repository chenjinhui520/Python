import re
import sys
from openpyxl import load_workbook


def get_biz_data(biz):
    data = load_workbook('/usr/local/conf/biz.xlsx')
    sheet = data['Sheet1']
    rows = sheet.rows
    biz_dict = {}
    for row in rows:
        line = [col.value for col in row]
        biz_dict[line[0]] = line[1:]

    if biz in biz_dict.keys():
        biz_id = str(biz_dict.get(biz)[0])
        department = str(biz_dict.get(biz)[1])
        department_id = str(biz_dict.get(biz)[2])
        return biz_id, department, department_id
    else:
        print('您输入的业务分类 P 不存在，请查证！')
        sys.exit()

def get_lb_name():
    lb_name = str(input('请输入负载均衡名称：')).strip().lower()
    lb_name = re.sub(r'\.', '-', lb_name)
    if len(lb_name) > 32:
        domain_name = str(input('请输入需要进行简称映射的域名：')).strip().lower()
        domain_name = re.sub(r'\.', '-', domain_name)

        data = load_workbook('/usr/local/conf/domain-mapping.xlsx')
        sheet = data['Sheet1']
        rows = sheet.rows
        domain_dict = {}
        for row in rows:
            line = [col.value for col in row]
            domain_dict[line[0]] = line[1]

        if domain_name in domain_dict.keys():
            domain_value = domain_dict[domain_name]
            lb_name = re.sub(domain_name, domain_value, lb_name)
            if len(lb_name) > 32:
                print('LB 名称替换后仍超过 32 个字符，系统退出！')
                sys.exit()
            else:
                return lb_name
        else:
            print('您输入的域名关系映射缺失，请查证！')
            sys.exit()
    else:
        return lb_name
