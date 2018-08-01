from openpyxl import load_workbook
from openpyxl.utils import get_column_letter

# 加载 exel 文件
data = load_workbook('C:\\Users\liygmf\Desktop\p2p.xlsx')

# 获得所有 sheet 名称
sheet_names = data.sheetnames

# 获得 sheet
sheet = data['p2p']

# 获取第二列，第四行的内容
# b4 = sheet.cell(row=4, column=2)
# print(b4.value)

# 获得最大行
max_row = sheet.max_row

# 获得最大列
max_column = sheet.max_column

# 获得最大列的字母号
max_letter = get_column_letter(sheet.max_column)

# 获得最小坐标
min_postion = 'A1'

# 获得最大坐标
max_postion = max_letter + str(max_row)

# 获得所有行，返回结果是生成器对象
rows = sheet.rows

# 获得所有列，返回结果是生成器对象
columns = sheet.columns

# 获得所有行的结果，第一行为字段名
for row in rows:
    line = [col.value for col in row]
    print(line)
